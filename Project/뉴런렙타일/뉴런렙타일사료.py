from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from operator import itemgetter
import urllib.request as req
import xlwings as xw
import requests
import time
import os


wb = Workbook()
ws = wb.active

i=1
option = webdriver.ChromeOptions()
option.add_argument('--headless')
img_list = []
text_list = []
price_list = []
href_list = []
while True:
    url = url = f'https://newrunreptile.co.kr/product/list.html?cate_no=355&page={i}'
    driver = webdriver.Chrome(options=option)
    driver.implicitly_wait(3)
    driver.get(url)
    time.sleep(1)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    driver.quit()

    div = soup.select_one("div.xans-product-listnormal")
    try :
        List = div.select_one("ul")
        lis = List.find_all("li")
        for li in lis:
            a_tag = li.select_one("a")
            if a_tag:
                # 하이퍼링크 가져오기
                a_href = a_tag.get("href")
                a_href = "https://raraaqua.com" + a_href
                href_list.append(a_href)
                print(a_href)
            img = li.select_one("img")
            if img:
                # 이미지 가져오기
                src = img.get('src')
                img_list.append(src)
                print("https:"+src)
            # 제목 가져오기
            p = li.select_one("p")
            if p:
                spans = p.find_all("span")
                print(spans[1].text)
                text_list.append(spans[1].text)
            # 가격 가져오기
            ul = li.select_one("ul")
            if ul:
                lis = ul.find_all("li")
                lspans = lis[1].find_all("span")
                print(lspans[1].text)
                price_list.append(lspans[1].text)
        print(f"{i}페이지끝")
    except Exception as e:
        print(str(e))
        break
    i+=1

n=1
for t, h, i, p in zip(text_list, href_list, img_list, price_list) :
    ws[f'A{n}'] = t
    ws[f'B{n}'] = h
    ws[f'C{n}'] = i
    ws[f'D{n}'] = p
    n+=1

wb.save("사료데이터(뉴런렙타일).xlsx")
wb.close()