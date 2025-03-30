from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
# openpyxl라이브러리에서 load_workbook을 import
from openpyxl import load_workbook
from operator import itemgetter
import urllib.request as req
import requests
import os
import requests
import json
import cx_Oracle

class DBManager:
    def __init__(self):
        self.conn = None

    def get_connection(self):
        try:
            if self.conn is None or self.conn.closed:
                self.conn = cx_Oracle.connect("Pro","pro","localhost:1521/xe")
            return self.conn
        except Exception as e:
            return None
    def __del__(self):
        """객체 소멸 시 연결 종료"""
        if self.conn:
            self.conn.close()
            print("db 연결이 정상적으로 종료되었습니다.")

    def insert(self, query, param):
        """데이터 삽입"""
        cursor = None
        try:
            if self.conn is None:
                self.get_connection()
            cursor = self.conn.cursor()
            cursor.execute(query, param)
            self.conn.commit()
        except Exception as e:
            if self.conn:
                self.conn.rollback()
        finally:
            if cursor:
                cursor.close()


if __name__ == '__main__':
    sql = """
            INSERT
            INTO fish_tank(title, img, price)
            VALUES(:1, :2, :3)
        """
    # '.xlsx' 엑셀 파일 불러오기
    wb = load_workbook(filename='어항데이터.xlsx')
    ws = wb['어항']
    n = 1

    db = DBManager()
    conn = db.get_connection()
    while True:
        if ws[f'A{n}'].value == None:
            break
        if conn:
            db.insert(sql, [ws[f'A{n}'].value, ws[f'B{n}'].value, ws[f'C{n}'].value])
        n+=1

            # 첫 번째 시트 불러오기
            # ws = wb.active



