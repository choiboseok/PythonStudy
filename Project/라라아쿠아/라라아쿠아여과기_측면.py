from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook
import time
import os

wb = Workbook()
ws = wb.active

option = webdriver.ChromeOptions()
option.add_argument('--headless')

url = 'https://raraaqua.com/category/%EC%B8%A1%EB%A9%B4%EC%97%AC%EA%B3%BC%EA%B8%B0/881/'
driver = webdriver.Chrome(options=option)
driver.implicitly_wait(3)
driver.get(url)
time.sleep(1)
soup = BeautifulSoup(driver.page_source, 'html.parser')
driver.quit()
img_list = []
text_list = []
price_list = []
href_list = []
div = soup.select_one("ul.prdList")
lis = div.find_all('li')
for li in lis:
    a_tag = li.select_one("a")
    if a_tag:
        a_href = a_tag.get("href")
        a_href = "https://raraaqua.com" + a_href
        href_list.append(a_href)
        print(a_href)
    img = li.select_one("img")
    if img:
        # 이미지, 제목 가져오기
        src = img.get('src')
        alt = img.get('alt')
        img_list.append(src)
        text_list.append(alt)
        print("https:"+src)
        print(alt)
    # 가격 가져오기
    description = li.select_one("div.description")
    if description:
        ul = description.select_one("ul")
        spans = ul.find_all('span')
        price = spans[1].text
        price_list.append(price)
        print(spans[1].text)


n=1
for t, h, i, p in zip(text_list, href_list, img_list, price_list) :
    ws[f'A{n}'] = t
    ws[f'B{n}'] = h
    ws[f'C{n}'] = i
    ws[f'D{n}'] = p
    n+=1

wb.save("/엑셀/라라/여과기데이터(라라아쿠아)_측면.xlsx")
wb.close()