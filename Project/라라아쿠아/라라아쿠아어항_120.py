from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook
import time
import os
img_path = '../testimg'
if not os.path.exists(img_path):
    os.mkdir(img_path)

# wb = Workbook()
# ws = wb.active


url = 'https://raraaqua.com/category/120cm-%EC%9D%B4%EC%83%81/624/'
driver = webdriver.Chrome()
driver.implicitly_wait(3)
driver.get(url)
time.sleep(1)
soup = BeautifulSoup(driver.page_source, 'html.parser')
driver.quit()
img_list = []
text_list = []
price_list = []
div = soup.select_one("ul.prdList")
lis = div.find_all('li')
for li in lis:
    img = li.select_one("img")
    if img:
        # 이미지, 제목 가져오기
        src = img.get('src')
        alt = img.get('alt')
        img_list.append(src)
        text_list.append(alt)
        print("https:"+src)
        print(alt)
    # 가격 가져오기
    description = li.select_one("div.description")
    if description:
        ul = description.select_one("ul")
        spans = ul.find_all('span')
        price = spans[1].text
        price_list.append(price)
        print(spans[1].text)


# n=1
# for t, i, p in zip(text_list, img_list, price_list) :
#     ws[f'A{n}'] = t
#     ws[f'B{n}'] = i
#     ws[f'C{n}'] = p
#     ws[f'D{n}'] = '어항'
#     n+=1

# wb.save("어항데이터(라라아쿠아).xlsx")
# wb.close()