from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from operator import itemgetter
import urllib.request as req
import xlwings as xw
import requests
import time
import os

img_path = '../week2/ex_crawling/testimg'
if not os.path.exists(img_path):
    os.mkdir(img_path)

wb = Workbook()
ws = wb.active

query = '세진아쿠아리움'
i=1
option = webdriver.ChromeOptions()
option.add_argument('--headless')
img_list = []
text_list = []
price_list = []
while True:
    url = url = f'https://seijin.co.kr/goods/goods_list.php?page={i}&cateCd=010'
    driver = webdriver.Chrome(options=option)
    driver.implicitly_wait(3)
    driver.get(url)
    time.sleep(1)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    driver.quit()

    div = soup.select_one("div.item_gallery_type")
    lis = div.find_all('li')
    if lis ==[]:
        print('수집종료')
        break
    for li in lis:
        # 사진 가져오기
        photo = li.select_one('.item_photo_box')
        p_a = photo.select_one('a')
        img = p_a.select_one('img')
        img_list.append(img['src'])
        # 타이틀 가져오기
        info = li.select_one('.item_info_cont')
        i_a = info.select_one('a')
        name = i_a.select_one('strong.item_name')
        text_list.append(name.text)
        # 가격 가져오기
        money_box = li.select_one('div.item_money_box')
        price = money_box.select_one('strong.item_price')
        price_list.append(price.text)
        print(img['src'])
        print(name.text)
        print(price.text)
        print("-"*100)
    print(i, "페이지 끝")
    # page 값 증가
    i += 1

ws.title = "히터"
n=1
for t, i, p in zip(text_list, img_list, price_list) :
    ws[f'A{n}'] = t
    ws[f'B{n}'] = i
    ws[f'C{n}'] = p
    n+=1

#
# n=1
# for t in text_set :
#     ws[f'A{n}'] = t
#     n+=1
#
# n=1
# for i in img_set :
#     ws[f'B{n}'] = i
#     n+=1
#
# n=1
# for i in price_set :
#     ws[f'C{n}'] = i
#     n+=1

# row = ['']
# ws.append(row)
# row = sorted(row, key=itemgetter(0))
# for data in row:
#     pass

wb.save("어항데이터.xlsx")
wb.close()